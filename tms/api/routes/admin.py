"""
System Admin routes
"""
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from typing import Dict, Any

from tms.infra.database import get_db
from tms.infra.models import User
from tms.api.dependencies import require_admin
from tms.api.schemas.common import SystemStats
from tms.application.services.system_service import SystemService

router = APIRouter(prefix="/admin", tags=["System Admin"])


@router.get("/stats", response_model=SystemStats, dependencies=[Depends(require_admin)])
def get_system_statistics(db: Session = Depends(get_db)):
    """Get system-wide statistics (Admins only)"""
    system_service = SystemService(db)
    stats = system_service.get_system_statistics()
    return SystemStats(**stats)


@router.get("/config", dependencies=[Depends(require_admin)])
def get_system_config(db: Session = Depends(get_db)):
    """Get system configuration (Admins only)"""
    system_service = SystemService(db)
    config = system_service.get_config()
    return config


@router.put("/config", dependencies=[Depends(require_admin)])
def update_system_config(
    config_updates: Dict[str, Any],
    db: Session = Depends(get_db)
):
    """Update system configuration (Admins only)"""
    system_service = SystemService(db)
    success = system_service.update_config(config_updates)
    
    if not success:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to update configuration"
        )
    
    return {"message": "Configuration updated successfully"}


@router.post("/users/{user_id}/freeze", dependencies=[Depends(require_admin)])
def freeze_user(user_id: int, db: Session = Depends(get_db)):
    """Freeze (deactivate) a user account (Admins only)"""
    system_service = SystemService(db)
    success = system_service.freeze_user(user_id)
    
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    return {"message": "User account frozen successfully"}


@router.post("/users/{user_id}/unfreeze", dependencies=[Depends(require_admin)])
def unfreeze_user(user_id: int, db: Session = Depends(get_db)):
    """Unfreeze (activate) a user account (Admins only)"""
    system_service = SystemService(db)
    success = system_service.unfreeze_user(user_id)
    
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    return {"message": "User account unfrozen successfully"}


@router.post("/backup", dependencies=[Depends(require_admin)])
def create_backup(db: Session = Depends(get_db)):
    """Create a database backup (Admins only)"""
    system_service = SystemService(db)
    backup_path = system_service.backup_database()
    return {"message": "Backup created successfully", "backup_path": backup_path}


@router.post("/semester", dependencies=[Depends(require_admin)])
def set_semester(semester: str, db: Session = Depends(get_db)):
    """Set the current semester (Admins only)"""
    system_service = SystemService(db)
    success = system_service.set_current_semester(semester)
    
    if not success:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to set semester"
        )
    
    return {"message": f"Current semester set to {semester}"}


@router.post("/enrollment/toggle", dependencies=[Depends(require_admin)])
def toggle_enrollment(open: bool, db: Session = Depends(get_db)):
    """Enable or disable enrollment (Admins only)"""
    system_service = SystemService(db)
    success = system_service.toggle_enrollment(open)
    
    if not success:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to toggle enrollment"
        )
    
    status_text = "enabled" if open else "disabled"
    return {"message": f"Enrollment {status_text} successfully"}


# -------------------------------------------------------------------------
# User Management Endpoints
# -------------------------------------------------------------------------

@router.get("/users")
def list_users(
    skip: int = 0,
    limit: int = 100,
    role: str = None,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin)
):
    """List users with optional role filter"""
    from tms.infra.models import User, UserRole
    query = db.query(User)
    if role:
        query = query.filter(User.role == role)
    users = query.offset(skip).limit(limit).all()
    # Simple serialization to avoid massive schemas for now
    return [
        {
            "id": u.id,
            "username": u.username,
            "email": u.email,
            "full_name": u.full_name,
            "role": u.role.value,
            "is_active": u.is_active,
            "created_at": u.created_at
        }
        for u in users
    ]

@router.get("/users/template")
def get_user_template():
    """Download Excel template for user batch import"""
    from openpyxl import Workbook
    from fastapi.responses import Response
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "User Import Template"
    
    # Headers
    headers = ["Username", "Email", "Password", "Full Name", "Role (student/teacher/admin)"]
    ws.append(headers)
    
    # Sample data
    ws.append(["new_user", "user@example.com", "password123", "New User", "student"])
    
    # Adjust column widths
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="user_import_template.xlsx"'
    }
    return Response(content=output.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)

@router.post("/users/batch")
async def batch_import_users(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin)
):
    """Batch import users from Excel"""
    # Move imports to top or keep here if circular dependency risk? 
    # Usually safer to import services at method level if circular deps exist, 
    # but AuthService shouldn't depend on Admin router.
    # We will keep them here for safety against circular deps but wrap in try catch for import errors.
    
    try:
        from io import BytesIO
        from openpyxl import load_workbook
        from tms.application.services.auth_service import AuthService
        from tms.infra.models import UserRole
        
        if not file.filename.endswith('.xlsx'):
            raise HTTPException(status_code=400, detail="Invalid file format. Please upload .xlsx file")
            
        try:
            contents = await file.read()
            wb = load_workbook(filename=BytesIO(contents))
            ws = wb.active
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")
        
        auth_service = AuthService(db)
        results = {"success": 0, "failed": 0, "errors": []}
        
        # Skip header row
        rows = list(ws.rows)
        if len(rows) < 2:
             return results # No data
             
        for i, row in enumerate(rows[1:], start=2):
            try:
                # Extract data safely
                username = str(row[0].value).strip() if row[0].value else None
                email = str(row[1].value).strip() if row[1].value else None
                password = str(row[2].value).strip() if row[2].value else "password123" # Default if missing
                full_name = str(row[3].value).strip() if row[3].value else None
                role_str = str(row[4].value).strip().lower() if row[4].value else "student"
                
                if not username or not email:
                    continue  # Skip empty rows
                    
                # Map role
                role_map = {
                    "student": UserRole.STUDENT,
                    "teacher": UserRole.TEACHER,
                    "admin": UserRole.ADMIN
                }
                role = role_map.get(role_str, UserRole.STUDENT)
                
                # Create user
                user = auth_service.create_user(
                    username=username,
                    email=email,
                    password=password,
                    full_name=full_name or username,
                    role=role
                )
                
                if user:
                    results["success"] += 1
                else:
                    results["failed"] += 1
                    results["errors"].append(f"Row {i}: Username or email already exists")
                    
            except Exception as e:
                import traceback
                traceback.print_exc()
                results["failed"] += 1
                results["errors"].append(f"Row {i}: {str(e)}")
                
        return results

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Server error during import: {str(e)}")


@router.put("/users/{user_id}", dependencies=[Depends(require_admin)])
def update_user(
    user_id: int,
    user_update: Dict[str, Any],
    db: Session = Depends(get_db)
):
    """
    Update user information (Admin only)
    Body should be a JSON object with optional fields:
    - username (str)
    - email (str)
    - full_name (str)
    - password (str)
    - role (str) - student/teacher/admin
    - is_active (bool)
    """
    from tms.application.services.auth_service import AuthService
    from tms.infra.models import UserRole
    from fastapi import HTTPException
    
    auth_service = AuthService(db)
    
    # Extract fields
    username = user_update.get("username")
    email = user_update.get("email")
    full_name = user_update.get("full_name")
    password = user_update.get("password")
    role_str = user_update.get("role")
    is_active = user_update.get("is_active")
    
    # Map role string to Enum if present
    role = None
    if role_str:
        try:
            role = UserRole(role_str.lower())
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid role")
            
    updated_user = auth_service.update_user(
        user_id=user_id,
        username=username,
        email=email,
        full_name=full_name,
        password=password,
        role=role,
        is_active=is_active
    )
    
    if not updated_user:
        raise HTTPException(
            status_code=400, 
            detail="Update failed. User not found or username/email already exists."
        )
        
    return {
        "id": updated_user.id,
        "username": updated_user.username,
        "email": updated_user.email,
        "full_name": updated_user.full_name,
        "role": updated_user.role.value,
        "is_active": updated_user.is_active
    }


# -------------------------------------------------------------------------
# Course Management Endpoints
# -------------------------------------------------------------------------

@router.get("/courses")
def list_courses_admin(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin)
):
    """List all courses"""
    from tms.infra.models import Course
    courses = db.query(Course).offset(skip).limit(limit).all()
    
    return [
        {
            "id": c.id,
            "code": c.course_code,
            "name": c.name,
            "credits": c.credits,
            "semester": c.semester,
            "teacher": c.teacher.user.full_name if c.teacher and c.teacher.user else "Unknown",
            "is_active": c.is_active
        }
        for c in courses
    ]

@router.get("/courses/template")
def get_course_template():
    """Download Excel template for course creation"""
    from openpyxl import Workbook
    from fastapi.responses import Response
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Course Import Template"
    
    # Headers
    headers = ["Course Code", "Course Name", "Credits", "Description", "Semester", "Teacher Username (Optional)"]
    ws.append(headers)
    
    # Sample data
    ws.append(["CS102", "Advanced Programming", 3, "Advanced concepts in programming", "2024 Spring", "prof_zhang"])
    
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="course_import_template.xlsx"'
    }
    return Response(content=output.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)

@router.post("/courses/batch")
async def batch_import_courses(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin)
):
    """Batch import courses from Excel"""
    from openpyxl import load_workbook
    from tms.application.services.course_service import CourseService
    from tms.infra.repositories.user_repository import UserRepository
    from tms.infra.models import UserRole
    
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload .xlsx file")
        
    contents = await file.read()
    wb = load_workbook(filename=BytesIO(contents))
    ws = wb.active
    
    course_service = CourseService(db)
    user_repo = UserRepository(db)
    
    results = {"success": 0, "failed": 0, "errors": []}
    
    # Find a default teacher (fallback)
    default_teacher = None
    default_teacher_user = user_repo.db.query(User).filter(User.role == UserRole.TEACHER).first()
    if default_teacher_user: # This is a User object
         # We need the Teacher object associated with this user, but CourseService uses teacher_id (which usually maps to teacher table ID, not user ID, or depends on data model). 
         # Checking CourseService.create_course: input is teacher_id.
         # Checking models.py for Teacher...
         from tms.infra.models import Teacher
         default_teacher = db.query(Teacher).join(User).filter(User.role == UserRole.TEACHER).first()

    
    rows = list(ws.rows)
    for i, row in enumerate(rows[1:], start=2):
        try:
            code = str(row[0].value).strip() if row[0].value else None
            name = str(row[1].value).strip() if row[1].value else None
            credits = float(row[2].value) if row[2].value else 3.0
            desc = str(row[3].value).strip() if row[3].value else ""
            semester = str(row[4].value).strip() if row[4].value else "2024 Fall"
            teacher_username = str(row[5].value).strip() if row[5].value else None
            
            if not code or not name:
                continue

            # Determine teacher
            teacher_id = None
            if teacher_username:
                # Find teacher by username
                from tms.infra.models import Teacher
                t_user = user_repo.get_by_username(teacher_username)
                if t_user and t_user.role == UserRole.TEACHER:
                    # Find associated teacher record
                    teacher_obj = db.query(Teacher).filter(Teacher.user_id == t_user.id).first()
                    if teacher_obj:
                        teacher_id = teacher_obj.id
            
            if not teacher_id and default_teacher:
                teacher_id = default_teacher.id
            
            if not teacher_id:
                results["failed"] += 1
                results["errors"].append(f"Row {i}: No valid teacher found")
                continue

            course = course_service.create_course(
                course_code=code,
                name=name,
                teacher_id=teacher_id,
                credits=credits,
                capacity=100, # Default capacity
                semester=semester,
                description=desc
            )
            
            if course:
                results["success"] += 1
            else:
                results["failed"] += 1
                results["errors"].append(f"Row {i}: Course creation failed (maybe code exists)")
                
        except Exception as e:
            results["failed"] += 1
            results["errors"].append(f"Row {i}: {str(e)}")
            
    return results
