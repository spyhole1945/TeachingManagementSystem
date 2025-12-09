"""
Grade routes
"""
from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from typing import List

from tms.infra.database import get_db
from tms.infra.models import User
from tms.api.dependencies import get_current_user, require_teacher
from tms.api.schemas.common import GradeCreate, GradeResponse
from tms.application.services.grade_service import GradeService

router = APIRouter(prefix="/grades", tags=["Grades"])


@router.post("/", response_model=GradeResponse, dependencies=[Depends(require_teacher)])
def record_grade(
    grade: GradeCreate,
    db: Session = Depends(get_db)
):
    """Record or update a grade (Teachers and Admins only)"""
    grade_service = GradeService(db)
    
    recorded_grade = grade_service.record_grade(
        student_id=grade.student_id,
        course_id=grade.course_id,
        score=grade.score,
        letter_grade=grade.letter_grade,
        comments=grade.comments
    )
    
    if not recorded_grade:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid student, course, or score"
        )
    
    return GradeResponse.model_validate(recorded_grade)


@router.get("/student/{student_id}", response_model=List[GradeResponse])
def get_student_grades(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all grades for a student"""
    grade_service = GradeService(db)
    grades = grade_service.get_student_grades(student_id)
    return [GradeResponse.model_validate(g) for g in grades]


@router.get("/course/{course_id}", response_model=List[GradeResponse])
def get_course_grades(
    course_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all grades for a course"""
    grade_service = GradeService(db)
    grades = grade_service.get_course_grades(course_id)
    return [GradeResponse.model_validate(g) for g in grades]


@router.get("/course/{course_id}/statistics")
def get_course_statistics(
    course_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get statistical information for a course"""
    grade_service = GradeService(db)
    stats = grade_service.get_course_statistics(course_id)
    return stats


from fastapi import UploadFile, File
from fastapi.responses import StreamingResponse
import io
import openpyxl
from tms.infra.models import Student

@router.get("/template")
def get_grade_template(
    course_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Download grade entry template"""
    from tms.application.services.course_service import CourseService
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grades"
    
    # Headers
    headers = ["Student Number", "Student Name", "Score", "Comments"]
    ws.append(headers)
    
    # Pre-fill data if course_id provided
    if course_id:
        course_service = CourseService(db)
        students = course_service.get_enrolled_students(course_id)
        for student in students:
            # Student Number, Name, Blank Score, Blank Comments
            student_name = "Unknown"
            try:
                if student.user:
                    student_name = student.user.full_name
            except Exception:
                pass
            ws.append([student.student_number, student_name, "", ""])
    else:
        # Sample row if no course specified
        ws.append(["S2024001", "Sample Student", "", ""])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = "grade_template.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        output, 
        headers=headers, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@router.post("/batch/{course_id}", dependencies=[Depends(require_teacher)])
async def batch_upload_grades(
    course_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Batch upload grades via Excel"""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload .xlsx file")
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    from tms.application.services.notification_service import NotificationService
    from tms.infra.models import Course
    
    grade_service = GradeService(db)
    notification_service = NotificationService(db)
    
    # Get course info for notifications
    course = db.query(Course).filter(Course.id == course_id).first()
    course_name = course.name if course else "Unknown Course"
    
    success_count = 0
    failure_count = 0
    errors = []
    
    # Iterate over rows (min_row=2 to skip header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:  # Skip empty rows or rows without student number
            continue
            
        student_number = str(row[0]).strip()
        # Row structure is now: [Number, Name, Score, Comments]
        # So Score is index 2, Comments is index 3
        score = row[2] 
        comments = str(row[3]) if len(row) > 3 and row[3] else None
        
        try:
            # Validate score
            if score is None:
                raise ValueError("Score is missing")
            
            try:
                score_val = float(score)
            except ValueError:
                raise ValueError(f"Invalid score: {score}")
                
            # Find student
            student = db.query(Student).filter(Student.student_number == student_number).first()
            if not student:
                raise ValueError(f"Student not found: {student_number}")
                
            # Record grade
            grade_service.record_grade(
                student_id=student.id,
                course_id=course_id,
                score=score_val,
                comments=comments
            )
            
            # Send notification
            # Use user_id if student has user associated
            if student.user_id:
                notification_service.notify_grade_released(
                    student_id=student.user_id,
                    course_name=course_name,
                    score=score_val
                )
            
            success_count += 1
            
        except Exception as e:
            failure_count += 1
            errors.append(f"Row {row_idx}: {str(e)}")
    
    return {
        "success_count": success_count,
        "failure_count": failure_count,
        "errors": errors
    }
